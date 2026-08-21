"""
core.py

Pure data-processing logic for the AMBR plate-mapping tool, extracted from
generate_plate_mapping.py with all tkinter/CLI/dialog code removed. This
module has no GUI and no argparse -- it's imported by app.py (the Flask web
front-end) so the exact same parsing/merging/repacking logic that was
validated in the desktop script is reused unchanged for the web version.

See the desktop script's docstring (generate_plate_mapping.py) for the full
design rationale on: why Plate/Well always comes from the AMBR file's literal
text rather than the scheme grid, why samples not in the scheme are flagged
but never dropped, why plate format (24 vs 96-well) can only be confirmed by
the user rather than proven from well usage alone, and why transposition
pools every source plate into one continuous run before repacking.
"""

import csv
import os
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill

REACTOR_RE = re.compile(r"Bioreactor\s+(\d+)", re.IGNORECASE)
EVENT_RE = re.compile(
    r"Sample\s+([\d.]+)\s*mL\s+to\s+([A-Za-z0-9 _#]+?)\s*/\s*([A-H]\d{1,2})",
    re.IGNORECASE,
)
SAMPLE_ID_RE = re.compile(r"R(\d+)S(\d+)", re.IGNORECASE)
WELL_RE = re.compile(r"^([A-H])(\d{1,2})$", re.IGNORECASE)

PLATE_DIMENSIONS = {
    "24": (["A", "B", "C", "D"], 6),
    "96": (["A", "B", "C", "D", "E", "F", "G", "H"], 12),
}


# --------------------------------------------------------------------------
# 1. Parse the AMBR timepoints.csv
# --------------------------------------------------------------------------
def parse_timepoints(path, exclude_text, exclude_volumes):
    events = defaultdict(list)
    dropped = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)  # skip header row
        for row in reader:
            if not row or not row[0].strip():
                continue
            batch_id, time_raw, sampling_event = row[0], row[1], row[2]
            m = REACTOR_RE.search(batch_id)
            if not m:
                continue
            reactor_num = m.group(1)

            if any(t.lower() in sampling_event.lower() for t in exclude_text):
                dropped.append((batch_id, time_raw, sampling_event, "excluded text match"))
                continue

            ev = EVENT_RE.search(sampling_event)
            if not ev:
                dropped.append((batch_id, time_raw, sampling_event, "could not parse event text"))
                continue
            volume, plate_name, well = ev.group(1), ev.group(2).strip(), ev.group(3)

            if volume in exclude_volumes:
                dropped.append((batch_id, time_raw, sampling_event, "excluded volume match"))
                continue

            plate_num_m = re.search(r"(\d+)\s*$", plate_name)
            plate_num = plate_num_m.group(1) if plate_num_m else plate_name

            events[reactor_num].append(
                {
                    "time_raw": time_raw,
                    "time_value": float(time_raw),
                    "volume": volume,
                    "plate": plate_num,
                    "plate_name_raw": plate_name,
                    "well": well,
                }
            )

    for reactor_num in events:
        events[reactor_num].sort(key=lambda e: e["time_value"])

    return events, dropped


# --------------------------------------------------------------------------
# 2. Parse the Benchling Timepoint-Sample.csv template
# --------------------------------------------------------------------------
def parse_benchling_template(path):
    reactor_info = {}
    experiment_name = None
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            parent_culture = (row.get("Parent Culture") or "").strip()
            medium = (row.get("Medium") or "").strip()
            exp = (row.get("Experiment") or "").strip()
            if exp:
                experiment_name = exp
            m = re.search(r"_R(\d+)_", parent_culture)
            if m and parent_culture:
                reactor_info[m.group(1)] = {"parent_culture": parent_culture, "medium": medium}
    return experiment_name, reactor_info


# --------------------------------------------------------------------------
# 3. Parse the sampling scheme xlsx ("List Format" sheet)
# --------------------------------------------------------------------------
def parse_scheme(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["List Format"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    ncols = ws.max_column

    reactor_order = []
    seen = set()
    for row in rows:
        for cell in row:
            if not cell or not isinstance(cell, str):
                continue
            m = SAMPLE_ID_RE.match(cell.strip())
            if m:
                r = m.group(1)
                if r not in seen:
                    seen.add(r)
                    reactor_order.append(r)
    num_reactors = len(reactor_order) or 1

    scheme = {}
    for col in range(ncols):
        col_values = [rows[r][col] if col < len(rows[r]) else None for r in range(len(rows))]
        base = None
        for r, val in enumerate(col_values):
            if val and isinstance(val, str) and val.strip().lower() != "empty":
                m = SAMPLE_ID_RE.match(val.strip())
                if m:
                    chunk = r // num_reactors
                    base = int(m.group(2)) - chunk
                    break
        if base is None:
            continue

        for r, val in enumerate(col_values):
            reactor_pos = r % num_reactors
            chunk = r // num_reactors
            s_index = base + chunk
            if reactor_pos >= len(reactor_order):
                continue
            reactor_num = reactor_order[reactor_pos]
            if val and isinstance(val, str) and val.strip().lower() != "empty":
                scheme[(reactor_num, s_index)] = {"name": val.strip(), "in_scheme": True}
            elif (reactor_num, s_index) not in scheme:
                scheme[(reactor_num, s_index)] = {"name": "Empty", "in_scheme": False}

    return scheme


def derive_scheme_from_events(events):
    scheme = {}
    for reactor_num, evs in events.items():
        for s_index, ev in enumerate(evs):
            scheme[(reactor_num, s_index)] = {"name": f"R{reactor_num}S{s_index}", "in_scheme": True}
    return scheme


def write_scheme_workbook(events, out_path):
    plate_order = []
    for reactor_num in sorted(events.keys(), key=lambda x: int(x)):
        for ev in events[reactor_num]:
            if ev["plate"] not in plate_order:
                plate_order.append(ev["plate"])

    per_plate = {p: [] for p in plate_order}
    flat_list = []
    for reactor_num in sorted(events.keys(), key=lambda x: int(x)):
        for s_index, ev in enumerate(events[reactor_num]):
            sample_id = f"R{reactor_num}S{s_index}"
            per_plate[ev["plate"]].append((s_index, reactor_num, sample_id))
            flat_list.append((reactor_num, s_index, sample_id))

    for p in per_plate:
        per_plate[p].sort(key=lambda t: (t[0], int(t[1])))

    wb = openpyxl.Workbook()
    ws_list = wb.active
    ws_list.title = "List Format"
    ws_list.append([f"Plate {p}" for p in plate_order])
    max_len = max((len(v) for v in per_plate.values()), default=0)
    for i in range(max_len):
        row = []
        for p in plate_order:
            entries = per_plate[p]
            row.append(entries[i][2] if i < len(entries) else "Empty")
        ws_list.append(row)

    ws_flat = wb.create_sheet("1Col_List")
    for reactor_num, s_index, sample_id in sorted(flat_list, key=lambda t: (int(t[0]), t[1])):
        ws_flat.append([sample_id])

    bold = Font(bold=True)
    for c in range(1, len(plate_order) + 1):
        ws_list.cell(row=1, column=c).font = bold

    wb.save(out_path)


# --------------------------------------------------------------------------
# 4. Build merged rows
# --------------------------------------------------------------------------
def build_rows(events, scheme, experiment_name, reactor_info, scheme_source="file"):
    rows = []
    well_number_counter = defaultdict(int)

    for reactor_num in sorted(events.keys(), key=lambda x: int(x)):
        for s_index, ev in enumerate(events[reactor_num]):
            key = (reactor_num, s_index)
            scheme_entry = scheme.get(key)
            if scheme_entry and scheme_entry["in_scheme"]:
                sample_name = scheme_entry["name"]
                in_scheme = "YES (auto-generated)" if scheme_source == "auto" else "YES"
            else:
                sample_name = f"R{reactor_num}S{s_index}"
                in_scheme = "NO - not in scheme"

            well_number_counter[(reactor_num, ev["plate"])] += 1

            info = reactor_info.get(reactor_num, {})
            reactor_label = f"R{reactor_num}"
            timepoint_num = f"S{s_index:02d}"
            replicate_num = 1
            benchling_sample = f"{experiment_name}_{reactor_label}__{timepoint_num}"
            soa_sample = f"{benchling_sample}_SOA_#{replicate_num}"
            rows.append(
                {
                    "sample": sample_name,
                    "plate": ev["plate"],
                    "well": ev["well"],
                    "plate_name_raw": ev["plate_name_raw"],
                    "reactor_plate_position": f"{ev['plate_name_raw']}/{ev['well']}",
                    "reactor": reactor_label,
                    "timepoint_num": timepoint_num,
                    "timepoint_h_raw": ev["time_raw"],
                    "timepoint_h_value": ev["time_value"],
                    "volume": ev["volume"],
                    "well_number": well_number_counter[(reactor_num, ev["plate"])],
                    "parent_culture": info.get("parent_culture", ""),
                    "medium": info.get("medium", ""),
                    "in_scheme": in_scheme,
                    "replicate_num": replicate_num,
                    "benchling_sample": benchling_sample,
                    "soa_sample": soa_sample,
                }
            )

    used_keys = {(r, s) for r, evs in events.items() for s in range(len(evs))}
    unused = [
        (reactor, s_index)
        for (reactor, s_index), entry in scheme.items()
        if not entry["in_scheme"] and (reactor, s_index) not in used_keys
    ]
    return rows, sorted(unused, key=lambda k: (k[0], k[1]))


# --------------------------------------------------------------------------
# 5. Write the workbook
# --------------------------------------------------------------------------
def plate_order_from_rows(rows):
    order = []
    for row in rows:
        if row["plate"] not in order:
            order.append(row["plate"])
    return order


def write_workbook(rows, unused_slots, experiment_name, out_path):
    """Writes the two-sheet plate mapping output:

    - "Benchling": ready to re-upload to Benchling. Columns match the
      Benchling Timepoint-Sample import template exactly (Reactor/Plate
      Number, Timepoint (#), Timepoint (h), Volume, Volume Unit, Experiment,
      Parent Culture, Medium, Reactor/Plate Position) -- confirmed with the
      user against a real filled-in Benchling export, deliberately omitting
      "Entity" and "Control?" (present in that reference file, but the user
      asked to leave them out).
    - "Detailed info": every other computed field (SOA sample name,
      destination SOA plate, well numbering, in-scheme flag, etc.) for
      cross-checking -- this is the old "benchling" sheet, renamed.

    Rows not found in the supplied sampling scheme are still highlighted on
    both sheets (via row fill), even though the "Benchling" sheet itself has
    no room for an explicit "In Scheme?" column without breaking the
    upload-template column set.
    """
    wb = openpyxl.Workbook()

    ws_main = wb.active
    ws_main.title = "Benchling"
    main_header = [
        "Reactor/Plate Number", "Timepoint (#)", "Timepoint (h)", "Volume", "Volume Unit",
        "Experiment", "Parent Culture", "Medium", "Reactor/Plate Position",
    ]
    ws_main.append(main_header)
    for row in rows:
        ws_main.append(
            [
                row["reactor"], row["timepoint_num"], row["timepoint_h_value"], row["volume"], "mL",
                experiment_name, row["parent_culture"], row["medium"], row["reactor_plate_position"],
            ]
        )

    ws_b = wb.create_sheet("Detailed info")
    benchling_header = [
        "#", "Benchling sample", "Destination Plate", "Benchling sample SOA", "Sample", "Plate",
        "Destination Well", "Reactor/Plate Number", "Timepoint (#)", "Timepoint (h)", "Time_Value",
        "Volume", "Well_Number", "Parent culture", "Medium", "Dilution",
        "Raw Absorbance Value #1", "Raw Absorbance Value #2", "CDW", "replicate #",
        "In Scheme?", "Not used",
    ]
    ws_b.append(benchling_header)

    plate_order = plate_order_from_rows(rows)

    for i, row in enumerate(rows):
        dest_plate = next(
            (f"{experiment_name}_SOA_plate#{j + 1}" for j, p in enumerate(plate_order) if p == row["plate"]),
            "UNKNOWN PLATE",
        )
        ws_b.append(
            [
                i + 1, row["benchling_sample"], dest_plate, row["soa_sample"], row["sample"], row["plate"],
                row["well"], row["reactor"], row["timepoint_num"], row["timepoint_h_raw"],
                row["timepoint_h_value"], row["volume"], row["well_number"], row["parent_culture"],
                row["medium"], None, None, None, None, row["replicate_num"], row["in_scheme"], None,
            ]
        )

    for reactor_num, s_index in unused_slots:
        ws_b.append(
            [None, None, None, None, "Empty", None, None, f"R{reactor_num}", f"S{s_index:02d}",
             None, None, None, None, None, None, None, None, None, None, None, "scheme slot unused", None]
        )

    helper_col_u = 23
    helper_col_v = 24
    ws_b.cell(row=2, column=helper_col_u, value=experiment_name or "")
    ws_b.cell(row=2, column=helper_col_v, value="CAREFUL TO WRITE THE PLATE NR AS TEXT")
    ws_b.cell(row=3, column=helper_col_u, value="Plate name benchling")
    ws_b.cell(row=3, column=helper_col_v, value="Plate Nr (from AMBR)")
    for i, plate_num in enumerate(plate_order):
        r = 4 + i
        ws_b.cell(row=r, column=helper_col_u, value=f"{experiment_name}_SOA_plate#{i + 1}")
        ws_b.cell(row=r, column=helper_col_v, value=str(plate_num))

    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold = Font(bold=True)
    flag_col_benchling = benchling_header.index("In Scheme?") + 1
    for i, row in enumerate(rows):
        r = i + 2  # +1 for header, +1 for 1-indexing
        if row["in_scheme"] == "NO - not in scheme":
            for c in range(1, len(main_header) + 1):
                ws_main.cell(row=r, column=c).fill = fill
    for r in range(2, ws_b.max_row + 1):
        if ws_b.cell(row=r, column=flag_col_benchling).value == "NO - not in scheme":
            for c in range(1, 21):
                ws_b.cell(row=r, column=c).fill = fill
    for c in range(1, len(main_header) + 1):
        ws_main.cell(row=1, column=c).font = bold
    for c in range(1, len(benchling_header) + 1):
        ws_b.cell(row=1, column=c).font = bold

    wb.save(out_path)


# --------------------------------------------------------------------------
# 6. 24-well <-> 96-well transposition
# --------------------------------------------------------------------------
def well_to_rowcol(well):
    m = WELL_RE.match(well.strip())
    if not m:
        raise ValueError(f"'{well}' doesn't look like a well address (expected e.g. 'A1', 'H12')")
    return m.group(1).upper(), int(m.group(2))


def detect_plate_format(rows):
    rows_letters_96 = PLATE_DIMENSIONS["96"][0]
    per_plate_max = {}
    for row in rows:
        try:
            letter, col = well_to_rowcol(row["well"])
        except ValueError:
            continue
        cur = per_plate_max.get(row["plate"], ("A", 0))
        cur_letter, cur_col = cur
        new_letter = letter if rows_letters_96.index(letter) > rows_letters_96.index(cur_letter) else cur_letter
        new_col = max(cur_col, col)
        per_plate_max[row["plate"]] = (new_letter, new_col)

    overall_max_letter, overall_max_col = "A", 0
    for letter, col in per_plate_max.values():
        if rows_letters_96.index(letter) > rows_letters_96.index(overall_max_letter):
            overall_max_letter = letter
        overall_max_col = max(overall_max_col, col)

    is_96 = rows_letters_96.index(overall_max_letter) > 3 or overall_max_col > 6
    verdict = "96" if is_96 else "24"
    return {
        "format": verdict,
        "max_row_letter": overall_max_letter,
        "max_col": overall_max_col,
        "per_plate_max": per_plate_max,
    }


def build_transposition(rows, dst_fmt):
    dst_rows_letters, dst_cols = PLATE_DIMENSIONS[dst_fmt]
    dst_capacity = len(dst_rows_letters) * dst_cols

    plate_order = plate_order_from_rows(rows)
    ordered = []
    for plate in plate_order:
        same_plate = [r for r in rows if r["plate"] == plate]
        same_plate.sort(key=lambda r: well_to_rowcol(r["well"]))
        ordered.extend(same_plate)

    transposed = []
    for i, row in enumerate(ordered):
        dst_plate_index = i // dst_capacity + 1
        pos = i % dst_capacity
        dst_letter = dst_rows_letters[pos // dst_cols]
        dst_col = pos % dst_cols + 1
        dst_well = f"{dst_letter}{dst_col}"
        transposed.append(
            {
                **row,
                "source_plate": row["plate"],
                "source_well": row["well"],
                "dest_plate_number": dst_plate_index,
                "dest_plate_label": f"Plate {dst_plate_index}",
                "dest_well": dst_well,
                "sequence_number": i + 1,
            }
        )
    return transposed


def write_transposed_workbook(transposed_rows, src_fmt, dst_fmt, out_path):
    """Writes the HPLC/SOA repack file. First sheet ("repacking SOA") is a
    simple repacking map -- just enough columns to answer "where did this
    sample come from, and where does it go": Sample, Source Plate, Source
    Well, Destination Plate, Destination Well, Benchling sample SOA Name, in
    that order (per the user's explicit column spec). Second sheet
    (plate_lookup) is unchanged -- same one-row-per-sample lookup, sorted by
    timepoint then reactor.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "repacking SOA"
    header = [
        "Sample", "Source Plate", "Source Well", "Destination Plate", "Destination Well",
        "Benchling sample SOA Name",
    ]
    ws.append(header)
    for row in transposed_rows:
        ws.append(
            [
                row["sample"], row["source_plate"], row["source_well"],
                row["dest_plate_label"], row["dest_well"], row["soa_sample"],
            ]
        )
    bold = Font(bold=True)
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).font = bold
    ws.freeze_panes = "A2"

    ws_lookup = wb.create_sheet("plate_lookup")
    ws_lookup.append(["Benchling sample SOA", "Destination Plate", "Destination Well"])
    for row in sorted(transposed_rows, key=lambda r: (int(r["timepoint_num"][1:]), int(r["reactor"][1:]))):
        ws_lookup.append([row["soa_sample"], row["dest_plate_label"], row["dest_well"]])
    for c in range(1, 4):
        ws_lookup.cell(row=1, column=c).font = bold
    ws_lookup.freeze_panes = "A2"

    wb.save(out_path)


def scan_volumes(path):
    counts = defaultdict(int)
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if not row or len(row) < 3 or not row[0].strip():
                continue
            ev = EVENT_RE.search(row[2])
            if ev:
                counts[ev.group(1)] += 1
    return counts
